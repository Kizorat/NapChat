#!/usr/bin/env python3
"""
Ponte Excel <-> CSV/JSON per il golden set delle traduzioni napoletane.

Serve a lavorare il golden set a mano: si esporta un foglio Excel formattato,
si rivede/corregge la colonna `napoletano` (e si annota in `note`), poi si
reimporta rigenerando il .csv e il .json ESISTENTI.

    export : golden_validator.json  ->  golden_validator.xlsx
    import : golden_validator.xlsx  ->  golden_validator.csv + golden_validator.json
    verify : confronta xlsx e json campo per campo (esce 1 se divergono)

Struttura del workbook: UN SOLO foglio "Golden" con tutti i turni di tutte le
conversazioni, una riga per turno, nello stesso ordine del JSON:

    conversazione | regione | macro_regione | languages |
    turn_index | tu_id | speaker | italiano | napoletano | note

Con --split si ottiene invece la variante a piu' fogli (un foglio "Conversazioni"
con i metadati + un foglio per conversazione); l'import riconosce entrambi i
formati automaticamente.

L'italiano viene dal corpus e non va toccato: in export le sue colonne sono
bloccate (foglio protetto, solo `napoletano` e `note` restano scrivibili) e in
import ogni divergenza rispetto al dataset di origine fa fallire il comando,
a meno di passare --allow-source-edits.

Prerequisiti:
    pip install openpyxl

Uso tipico:
    python scripts/golden_workbook.py export
    #  ... revisione manuale in Excel ...
    python scripts/golden_workbook.py import
    python scripts/golden_workbook.py verify
"""

import argparse, csv, json, os, sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# serializzatore condiviso col build script (stesso formato .json), che sta in
# scripts/ alla radice: questo file e' in golden_translate/script/, quindi il
# path va risalito fino alla radice del progetto.
ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, os.path.join(ROOT, "scripts"))
from build_conversation_dataset import write_json

GOLDEN_DIR = "golden_translate"
DEFAULT_JSON = os.path.join(GOLDEN_DIR, "golden_validator.json")
DEFAULT_XLSX = os.path.join(GOLDEN_DIR, "golden_validator.xlsx")
DEFAULT_CSV = os.path.join(GOLDEN_DIR, "golden_validator.csv")
DEFAULT_SOURCE = os.path.join("dataset_filter", "dataset_structured.json")

MAIN_SHEET = "Golden"                 # foglio unico (formato predefinito)
META_SHEET = "Conversazioni"          # foglio metadati (solo con --split)

META_FIELDS = ["id", "regione", "macro_regione", "languages", "n_turni_italiani"]
CONV_FIELDS = ["conversazione", "regione", "macro_regione", "languages"]
TURN_FIELDS = ["turn_index", "tu_id", "speaker", "italiano", "napoletano", "note"]
ALL_FIELDS = CONV_FIELDS + TURN_FIELDS          # colonne del foglio unico
# colonne che replicano il corpus e non vanno modificate a mano
LOCKED_FIELDS = set(CONV_FIELDS) | {"turn_index", "tu_id", "speaker", "italiano"}
CSV_FIELDS = ALL_FIELDS

COL_WIDTHS = {"conversazione": 14, "regione": 16, "macro_regione": 14, "languages": 12,
              "turn_index": 11, "tu_id": 9, "speaker": 11,
              "italiano": 60, "napoletano": 60, "note": 30}

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
LOCKED_FILL = PatternFill("solid", fgColor="F2F2F2")   # colonne di sola lettura
EDIT_FILL = PatternFill("solid", fgColor="FFF9E6")     # colonne da compilare
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_samples(path):
    """Carica il dataset in qualsiasi forma: oggetto JSON, array JSON o JSONL."""
    raw = open(path, encoding="utf-8").read().strip()
    try:
        obj = json.loads(raw)
        return obj if isinstance(obj, list) else [obj]
    except json.JSONDecodeError:
        return [json.loads(l) for l in raw.splitlines() if l.strip()]


def flatten(samples):
    """Dal JSON alla lista piatta di righe (una per turno), nell'ordine del JSON."""
    rows = []
    for s in samples:
        for t in s["turni"]:
            rows.append({
                "conversazione": s["id"], "regione": s.get("regione", ""),
                "macro_regione": s.get("macro_regione", ""),
                "languages": s.get("languages", ""),
                "turn_index": t["turn_index"], "tu_id": t.get("tu_id", ""),
                "speaker": t["speaker"], "italiano": t["italiano"],
                "napoletano": t.get("napoletano", ""), "note": t.get("note", ""),
            })
    return rows


# ----------------------------------------------------------------- export ----

def set_cell(ws, row, col, value):
    """Scrive una cella preservando il testo alla lettera.

    Il napoletano e' pieno di apostrofi iniziali ('o, 'a, 'e, 'nfatti): Excel li
    interpreta come marcatore di testo e li mangia appena l'utente ricommette la
    cella. Con quotePrefix l'apostrofo resta parte del valore.
    """
    c = ws.cell(row=row, column=col, value=value)
    if isinstance(value, str) and value.startswith("'"):
        c.quotePrefix = True
    return c


def style_header(ws, fields):
    for col, name in enumerate(fields, 1):
        c = ws.cell(row=1, column=col, value=name)
        c.fill, c.font, c.border = HEADER_FILL, HEADER_FONT, BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(fields))}1"
    ws.row_dimensions[1].height = 24


def protect(ws):
    """Blocca il foglio lasciando filtro, ordinamento e ridimensionamento colonne."""
    ws.protection.sheet = True
    ws.protection.autoFilter = False
    ws.protection.sort = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False


def add_nap_hint(ws, fields, n_rows):
    """Suggerimento sulla colonna napoletano (e segnala le celle lasciate vuote)."""
    col = get_column_letter(fields.index("napoletano") + 1)
    dv = DataValidation(type="textLength", operator="greaterThan", formula1="0",
                        allow_blank=False, showErrorMessage=False,
                        showInputMessage=True,
                        promptTitle="Traduzione napoletana",
                        prompt="Testo napoletano del turno. Non lasciare vuoto.")
    ws.add_data_validation(dv)
    dv.add(f"{col}2:{col}{n_rows + 1}")


def write_rows(ws, fields, rows):
    """Riempie un foglio con le righe date, applicando stile e blocchi."""
    style_header(ws, fields)
    for col, name in enumerate(fields, 1):
        ws.column_dimensions[get_column_letter(col)].width = COL_WIDTHS[name]

    for row, r in enumerate(rows, 2):
        for col, name in enumerate(fields, 1):
            c = set_cell(ws, row, col, r.get(name, ""))
            c.border = BORDER
            locked = name in LOCKED_FIELDS
            c.fill = LOCKED_FILL if locked else EDIT_FILL
            c.protection = Protection(locked=locked)
            if name in ("italiano", "napoletano", "note"):
                c.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                c.alignment = Alignment(horizontal="center", vertical="top")

    add_nap_hint(ws, fields, len(rows))
    protect(ws)


def write_meta_sheet(ws, samples):
    style_header(ws, META_FIELDS)
    widths = {"id": 12, "regione": 20, "macro_regione": 16,
              "languages": 20, "n_turni_italiani": 18}
    for col, name in enumerate(META_FIELDS, 1):
        ws.column_dimensions[get_column_letter(col)].width = widths[name]
    for row, s in enumerate(samples, 2):
        for col, name in enumerate(META_FIELDS, 1):
            c = set_cell(ws, row, col, s.get(name, ""))
            c.border, c.fill = BORDER, LOCKED_FILL
            c.alignment = Alignment(horizontal="center", vertical="center")
    protect(ws)


def cmd_export(args):
    samples = load_samples(args.json)
    rows = flatten(samples)
    os.makedirs(os.path.dirname(args.xlsx) or ".", exist_ok=True)

    wb = Workbook()
    if args.split:
        write_meta_sheet(wb.active, samples)
        wb.active.title = META_SHEET
        for s in samples:
            conv_rows = [r for r in rows if r["conversazione"] == s["id"]]
            write_rows(wb.create_sheet(title=s["id"][:31]), TURN_FIELDS, conv_rows)
        fogli = f"{META_SHEET} + " + ", ".join(s["id"] for s in samples)
    else:
        ws = wb.active
        ws.title = MAIN_SHEET
        write_rows(ws, ALL_FIELDS, rows)
        fogli = MAIN_SHEET

    wb.save(args.xlsx)

    print(f"Letto  : {args.json}")
    print(f"Scritto: {args.xlsx}")
    print(f"Fogli  : {fogli}")
    print(f"Righe  : {len(rows)} turni  ("
          + ", ".join(f"{s['id']} {len(s['turni'])}" for s in samples) + ")")
    print("\nColonne grigie = dal corpus (bloccate). Colonne gialle = napoletano, note.")


# ----------------------------------------------------------------- import ----

def read_sheet_rows(ws, fields):
    """Legge un foglio come lista di dict, mappando le colonne dall'intestazione."""
    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(h).strip() if h is not None else "" for h in next(rows)]
    except StopIteration:
        return []
    missing = [f for f in fields if f not in header]
    if missing:
        raise SystemExit(f"ERRORE: nel foglio {ws.title!r} mancano le colonne: "
                         f"{', '.join(missing)}")
    idx = {f: header.index(f) for f in fields}
    out = []
    for r in rows:
        if r is None or all(v is None or str(v).strip() == "" for v in r):
            continue                                   # riga vuota -> ignorata
        out.append({f: r[idx[f]] for f in fields})
    return out


def as_text(v):
    """Normalizza il valore letto da Excel a stringa pulita."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


def read_flat_rows(xlsx):
    """Legge il workbook (foglio unico o a piu' fogli) come lista di righe piatte."""
    wb = load_workbook(xlsx, data_only=True)

    if MAIN_SHEET in wb.sheetnames:                    # formato predefinito
        return [{f: as_text(r[f]) for f in ALL_FIELDS}
                for r in read_sheet_rows(wb[MAIN_SHEET], ALL_FIELDS)]

    if META_SHEET in wb.sheetnames:                    # formato --split
        rows = []
        for m in read_sheet_rows(wb[META_SHEET], META_FIELDS):
            conv = as_text(m["id"])
            if conv not in wb.sheetnames:
                raise SystemExit(f"ERRORE: manca il foglio dei turni per {conv!r}")
            for r in read_sheet_rows(wb[conv], TURN_FIELDS):
                row = {"conversazione": conv, "regione": as_text(m["regione"]),
                       "macro_regione": as_text(m["macro_regione"]),
                       "languages": as_text(m["languages"])}
                row.update({f: as_text(r[f]) for f in TURN_FIELDS})
                rows.append(row)
        return rows

    raise SystemExit(f"ERRORE: in {xlsx!r} non trovo ne' il foglio {MAIN_SHEET!r} "
                     f"ne' {META_SHEET!r}. Rigenera il workbook con 'export'.")


def rows_to_samples(rows):
    """Dalle righe piatte ai campioni nella struttura del JSON, con i problemi trovati."""
    problemi, per_conv = [], {}

    for pos, r in enumerate(rows, 2):                  # 2 = prima riga dati in Excel
        conv = r["conversazione"]
        if not conv:
            problemi.append(f"riga {pos}: conversazione vuota")
            continue
        ti = r["turn_index"]
        if not ti.isdigit():
            problemi.append(f"riga {pos} ({conv}): turn_index non numerico ({ti!r})")
            continue
        ti = int(ti)
        d = per_conv.setdefault(conv, {"meta": r, "turni": {}})
        if ti in d["turni"]:
            problemi.append(f"{conv}: turn_index {ti} duplicato")
            continue
        if not r["napoletano"]:
            problemi.append(f"{conv} turno {ti}: napoletano vuoto")
        d["turni"][ti] = r

    samples, flat = [], []
    for conv in sorted(per_conv):
        d = per_conv[conv]
        turni = [d["turni"][k] for k in sorted(d["turni"])]
        if [t["turn_index"] and int(t["turn_index"]) for t in turni] != \
                list(range(1, len(turni) + 1)):
            problemi.append(f"{conv}: turn_index non contigui da 1 a {len(turni)}")
        m = d["meta"]
        samples.append({
            "id": conv, "regione": m["regione"], "macro_regione": m["macro_regione"],
            "languages": m["languages"], "n_turni_italiani": len(turni),
            "turni": [{"turn_index": int(t["turn_index"]), "tu_id": t["tu_id"],
                       "speaker": t["speaker"], "italiano": t["italiano"],
                       "napoletano": t["napoletano"]} for t in turni],
        })
        flat.extend(turni)
    return samples, flat, problemi


def check_corpus_drift(samples, source_path):
    """Verifica che le colonne prese dal corpus non siano state modificate a mano."""
    if not os.path.exists(source_path):
        print(f"! {source_path} non trovato: salto il confronto col corpus",
              file=sys.stderr)
        return []
    source = {s["id"]: s for s in load_samples(source_path)}
    drift = []
    for s in samples:
        src = source.get(s["id"])
        if src is None:
            continue                                   # conversazione non nel corpus
        by_idx = {t["turn_index"]: t for t in src["turni"]}
        if len(s["turni"]) != len(src["turni"]):
            drift.append(f"{s['id']}: {len(s['turni'])} turni nel workbook vs "
                         f"{len(src['turni'])} nel corpus")
        for t in s["turni"]:
            o = by_idx.get(t["turn_index"])
            if o is None:
                drift.append(f"{s['id']} turno {t['turn_index']}: assente nel corpus")
                continue
            for f in ("italiano", "speaker", "tu_id"):
                if t[f] != o[f]:
                    drift.append(f"{s['id']} turno {t['turn_index']}: {f} modificato "
                                 f"({o[f]!r} -> {t[f]!r})")
    return drift


def bail(titolo, voci):
    head = voci[:25]
    sys.exit(f"{titolo}\n  - " + "\n  - ".join(head)
             + (f"\n  ... e altri {len(voci) - len(head)}" if len(voci) > len(head) else ""))


def cmd_import(args):
    if not os.path.exists(args.xlsx):
        sys.exit(f"ERRORE: workbook non trovato: {args.xlsx!r}\n"
                 f"Generalo con:  python scripts/golden_workbook.py export")

    rows = read_flat_rows(args.xlsx)
    if not rows:
        sys.exit(f"ERRORE: nessuna riga di dati in {args.xlsx!r}")

    samples, flat, problemi = rows_to_samples(rows)
    if problemi:
        bail("IMPORT ANNULLATO, problemi nel workbook:", problemi)

    drift = check_corpus_drift(samples, args.source)
    if drift and not args.allow_source_edits:
        bail("IMPORT ANNULLATO: colonne dal corpus modificate nel workbook "
             "(ripristinale, oppure rilancia con --allow-source-edits):", drift)
    if drift:
        print(f"! {len(drift)} divergenze dal corpus accettate (--allow-source-edits)",
              file=sys.stderr)

    os.makedirs(os.path.dirname(args.json) or ".", exist_ok=True)
    with open(args.csv, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        w.writeheader()
        w.writerows(flat)
    write_json(samples, args.json)

    print(f"Letto  : {args.xlsx}")
    print(f"Scritto: {args.csv}")
    print(f"Scritto: {args.json}")
    for s in samples:
        note = sum(1 for r in flat
                   if r["conversazione"] == s["id"] and r["note"])
        print(f"  - {s['id']}: {s['n_turni_italiani']} turni tradotti"
              + (f", {note} con note" if note else ""))
    print(f"Turni totali: {sum(s['n_turni_italiani'] for s in samples)}")


# ----------------------------------------------------------------- verify ----

def cmd_verify(args):
    """Controlla che il contenuto dell'Excel sia identico a quello del JSON."""
    for p in (args.xlsx, args.json):
        if not os.path.exists(p):
            sys.exit(f"ERRORE: file non trovato: {p!r}")

    xl, _, problemi = rows_to_samples(read_flat_rows(args.xlsx))
    js = sorted(load_samples(args.json), key=lambda s: s["id"])
    diffs = list(problemi)

    ids_xl, ids_js = [s["id"] for s in xl], [s["id"] for s in js]
    if ids_xl != ids_js:
        diffs.append(f"conversazioni diverse: xlsx {ids_xl} vs json {ids_js}")
    else:
        for a, b in zip(xl, js):
            for f in META_FIELDS:
                va = a["id"] if f == "id" else a[f]
                if str(va) != str(b.get(f, "")):
                    diffs.append(f"{a['id']}: metadato {f} xlsx={va!r} json={b.get(f)!r}")
            if len(a["turni"]) != len(b["turni"]):
                diffs.append(f"{a['id']}: {len(a['turni'])} turni nell'xlsx "
                             f"vs {len(b['turni'])} nel json")
            for ta, tb in zip(a["turni"], b["turni"]):
                for f in ("turn_index", "tu_id", "speaker", "italiano", "napoletano"):
                    if str(ta[f]) != str(tb.get(f, "")):
                        diffs.append(f"{a['id']} turno {ta['turn_index']}: {f}\n"
                                     f"      xlsx: {ta[f]!r}\n      json: {tb.get(f)!r}")

    tot = sum(len(s["turni"]) for s in xl)
    if diffs:
        head = diffs[:20]
        print(f"DIVERGENZE: {len(diffs)}\n  - " + "\n  - ".join(head)
              + (f"\n  ... e altre {len(diffs) - len(head)}" if len(diffs) > len(head) else ""),
              file=sys.stderr)
        sys.exit(1)
    print(f"OK: {args.xlsx} e {args.json} hanno contenuto identico "
          f"({len(xl)} conversazioni, {tot} turni, tutti i campi confrontati).")


# ------------------------------------------------------------------- main ----

def main():
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[1],
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = ap.add_subparsers(dest="cmd", required=True)

    e = sub.add_parser("export", help="JSON -> Excel formattato")
    e.add_argument("--json", default=DEFAULT_JSON, help=f"default: {DEFAULT_JSON}")
    e.add_argument("--xlsx", default=DEFAULT_XLSX, help=f"default: {DEFAULT_XLSX}")
    e.add_argument("--split", action="store_true",
                   help="un foglio per conversazione + foglio metadati "
                        "(default: foglio unico con tutti i turni)")
    e.set_defaults(func=cmd_export)

    i = sub.add_parser("import", help="Excel -> CSV + JSON")
    i.add_argument("--xlsx", default=DEFAULT_XLSX, help=f"default: {DEFAULT_XLSX}")
    i.add_argument("--csv", default=DEFAULT_CSV, help=f"default: {DEFAULT_CSV}")
    i.add_argument("--json", default=DEFAULT_JSON, help=f"default: {DEFAULT_JSON}")
    i.add_argument("--source", default=DEFAULT_SOURCE,
                   help=f"dataset di origine per il confronto (default: {DEFAULT_SOURCE})")
    i.add_argument("--allow-source-edits", action="store_true",
                   help="accetta modifiche a italiano/speaker/tu_id")
    i.set_defaults(func=cmd_import)

    v = sub.add_parser("verify", help="controlla che Excel e JSON abbiano lo stesso contenuto")
    v.add_argument("--xlsx", default=DEFAULT_XLSX, help=f"default: {DEFAULT_XLSX}")
    v.add_argument("--json", default=DEFAULT_JSON, help=f"default: {DEFAULT_JSON}")
    v.set_defaults(func=cmd_verify)

    args = ap.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
